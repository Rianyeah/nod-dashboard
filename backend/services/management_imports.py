"""Allowlisted validate-preview-commit imports for dashboard-managed datasets."""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import zipfile
from calendar import month_name
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Mapping
from uuid import uuid4

from fastapi import HTTPException, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy import text

from database import async_session
from user_store import AppUser


MAX_FILE_BYTES = 8 * 1024 * 1024
MAX_TOTAL_BYTES = 20 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_PREVIEW_ROWS = 25
ALLOWED_EXTENSIONS = frozenset({".csv", ".xlsx"})
TARGETS = {
    "ticketing_swfm_non_inap": {
        "label": "Ticketing SWFM Non-INAP",
        "strategy": "upsert",
        "accepted_extensions": [".xlsx", ".csv"],
        "supports_multiple_files": True,
        "description": "PMS, PMG, FNA, dan BBM berdasarkan nomor ticket.",
    },
    "ticketing_fault_center": {
        "label": "Ticketing Fault Center",
        "strategy": "replace_period",
        "accepted_extensions": [".csv", ".xlsx"],
        "supports_multiple_files": False,
        "description": "Satu periode bulan diganti secara atomik.",
    },
}


FAULT_CSV_TO_COLUMN = {
    "Ticket Number Inap": "ticket_number_inap",
    "Ticket Number SWFM": "ticket_number_swfm",
    "Kategori TT": "kategori_tt",
    "Severity": "severity",
    "Type Ticket": "type_ticket",
    "Site Id": "site_id",
    "Site Name": "site_name",
    "Site Class": "site_class",
    "Cluster TO": "cluster_to",
    "Kabupaten Kota": "kabupaten_kota",
    "Impact": "impact",
    "Occured Time": "occured_time",
    "Created At": "created_at",
    "Tahun": "tahun",
    "Periode/Bulan": "periode_bulan",
    "Tanggal": "tanggal",
    "MTTR": "mttr",
    "Respon time": "respon_time",
    "TakeOver": "takeover",
    "PLN downtime": "pln_downtime",
    "Durasi": "durasi",
    "Visitation": "visitation",
    "Backup sukses": "backup_sukses",
    "Ticket Inap Status": "ticket_inap_status",
    "Ticket SWFM Status": "ticket_swfm_status",
    "PIC Take Over Ticket": "pic_take_over_ticket",
    "NOP": "nop",
    "Regional": "regional",
    "Area": "area",
    "Is Escalate": "is_escalate",
    "Escalate To": "escalate_to",
    "Cleared Time": "cleared_time",
    "Is Auto Resolved": "is_auto_resolved",
    "RH Start": "rh_start",
    "RH Start Time": "rh_start_time",
    "RH Stop": "rh_stop",
    "RH Stop Time": "rh_stop_time",
    "RC Owner": "rc_owner",
    "RC Category": "rc_category",
    "RC 1": "rc_1",
    "RC 2": "rc_2",
    "Note": "note",
    "Resolution Action": "resolution_action",
    "Take Over Date": "take_over_date",
    "Chek in At": "chek_in_at",
    "INAP RC 1": "inap_rc_1",
    "INAP RC 2": "inap_rc_2",
    "INAP Resolution Action": "inap_resolution_action",
    "SLA Status": "sla_status",
    "Fault Text": "fault_text",
    "NOSSA No": "nossa_no",
    "Assignee Group": "assignee_group",
    "Summary": "summary",
    "Description": "description",
    "Submitted Time": "submitted_time",
    "Incident Priority": "incident_priority",
    "Hub": "hub",
    "Is Excluded In KPI": "is_excluded_in_kpi",
    "Ticket Creation": "ticket_creation",
    "Ticket Creator": "ticket_creator",
    "Site Cleared On": "site_cleared_on",
    "Rank": "rank",
    "Closed at": "closed_at",
    "Follow up at": "follow_up_at",
    "Holding status": "holding_status",
}
FAULT_TIMESTAMPS = {
    "occured_time", "created_at", "cleared_time", "rh_start_time", "rh_stop_time",
    "take_over_date", "chek_in_at", "submitted_time", "site_cleared_on", "closed_at",
    "follow_up_at",
}
FAULT_INTERVALS = {"mttr", "respon_time"}
FAULT_INTEGERS = {"tahun", "tanggal", "rank"}
FAULT_DECIMALS = {"pln_downtime", "rh_start", "rh_stop"}
FAULT_BOOLEANS = {"is_escalate", "is_auto_resolved", "is_excluded_in_kpi"}
RAW_FAULT_RENAMED_HEADERS = {
    "Chek in At": "Check In At",
    "Fault Text": "Fault Level",
    "Closed at": "Closed At",
    "Follow up at": "Follow Up At",
    "Holding status": "Holding Status",
}
RAW_FAULT_DERIVED_HEADERS = {
    "Kategori TT", "Kabupaten Kota", "Tahun", "Periode/Bulan", "Tanggal", "MTTR",
    "Respon time", "TakeOver", "PLN downtime", "Durasi", "Visitation", "Backup sukses",
}
RAW_FAULT_REQUIRED_HEADERS = (
    set(FAULT_CSV_TO_COLUMN)
    .difference(RAW_FAULT_DERIVED_HEADERS)
    .difference(RAW_FAULT_RENAMED_HEADERS)
    .union(RAW_FAULT_RENAMED_HEADERS.values())
)
FAULT_MONTH_VARIANTS = (
    {"january", "januari"},
    {"february", "februari"},
    {"march", "maret"},
    {"april"},
    {"may", "mei"},
    {"june", "juni"},
    {"july", "juli"},
    {"august", "agustus"},
    {"september"},
    {"october", "oktober"},
    {"november"},
    {"december", "desember"},
)
FAULT_MISSING_LOCATION_VALUES = {"tidak ada", "tidak tersedia", "n/a", "na", "-"}


@dataclass
class ParsedRow:
    source_file: str
    source_row: int
    row_key: str | None
    payload: dict[str, object]
    errors: list[str] = field(default_factory=list)
    change_kind: str = "insert"


@dataclass
class ParsedImport:
    rows: list[ParsedRow]
    warnings: list[str]
    metadata: dict[str, object]


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    normalized = re.sub(r"\s+", " ", str(value).strip())
    return normalized or None


def normalize_pic_key(value: object) -> str | None:
    normalized = normalize_text(value)
    return normalized.casefold() if normalized else None


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").casefold()).strip()


def _first(row: dict[str, object], *candidates: str) -> object:
    normalized = {normalize_header(key): value for key, value in row.items()}
    for candidate in candidates:
        value = normalized.get(normalize_header(candidate))
        if value is not None and normalize_text(value) is not None:
            return value
    return None


def _parse_date(value: object) -> date | None:
    if value is None or normalize_text(value) is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    candidate = str(value).strip()
    for fmt in (
        "%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y", "%d/%m/%Y %H:%M",
        "%d-%m-%Y", "%m/%d/%Y",
    ):
        try:
            return datetime.strptime(candidate, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Format tanggal tidak dikenali: {candidate}")


def _json_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, timedelta):
        total_seconds = int(value.total_seconds())
        sign = "-" if total_seconds < 0 else ""
        total_seconds = abs(total_seconds)
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{sign}{hours}:{minutes:02d}:{seconds:02d}"
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _source_hash(payload: dict[str, object]) -> str:
    encoded = json.dumps(payload, sort_keys=True, ensure_ascii=False, separators=(",", ":"))
    return hashlib.sha256(encoded.encode("utf-8")).hexdigest()


def _inspect_xlsx_archive(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            entries = archive.infolist()
            uncompressed = sum(entry.file_size for entry in entries)
            compressed = max(1, sum(entry.compress_size for entry in entries))
            if uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES or uncompressed / compressed > 100:
                raise ValueError("Arsip XLSX terlalu besar setelah diekstrak")
            if not any(entry.filename == "[Content_Types].xml" for entry in entries):
                raise ValueError("Berkas bukan XLSX yang valid")
            if any(entry.filename.casefold().endswith("vbaproject.bin") for entry in entries):
                raise ValueError("Workbook dengan macro tidak diizinkan")
    except zipfile.BadZipFile as exc:
        raise ValueError("Berkas XLSX rusak atau tidak valid") from exc


def _rows_from_xlsx(content: bytes) -> tuple[list[str], list[dict[str, object]]]:
    _inspect_xlsx_archive(content)
    workbook = load_workbook(
        io.BytesIO(content),
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        worksheet = next((sheet for sheet in workbook.worksheets if sheet.max_row > 1), None)
        if worksheet is None:
            raise ValueError("Workbook tidak memiliki baris data")
        iterator = worksheet.iter_rows(values_only=True)
        headers = [normalize_text(value) or "" for value in next(iterator)]
        if not any(headers):
            raise ValueError("Header workbook kosong")
        rows = [dict(zip(headers, values, strict=False)) for values in iterator]
        return headers, rows
    finally:
        workbook.close()


def _rows_from_csv(content: bytes) -> tuple[list[str], list[dict[str, object]]]:
    try:
        decoded = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValueError("CSV harus menggunakan encoding UTF-8") from exc
    sample = decoded[:4096]
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=";,\t,").delimiter
    except csv.Error:
        delimiter = ";"
    reader = csv.DictReader(io.StringIO(decoded), delimiter=delimiter)
    headers = [normalize_text(value) or "" for value in (reader.fieldnames or [])]
    if not headers:
        raise ValueError("Header CSV kosong")
    return headers, [dict(row) for row in reader]


async def _read_upload(upload: UploadFile) -> tuple[str, bytes]:
    filename = Path(upload.filename or "upload").name
    extension = Path(filename).suffix.casefold()
    if extension not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Format {extension or '(tanpa ekstensi)'} tidak didukung")
    content = await upload.read(MAX_FILE_BYTES + 1)
    if len(content) > MAX_FILE_BYTES:
        raise ValueError(f"Ukuran {filename} melebihi 8 MB")
    if not content:
        raise ValueError(f"Berkas {filename} kosong")
    return filename, content


def _tabular_rows(filename: str, content: bytes) -> tuple[list[str], list[dict[str, object]]]:
    return _rows_from_xlsx(content) if Path(filename).suffix.casefold() == ".xlsx" else _rows_from_csv(content)


def _classify_ticket_type(filename: str, rows: list[dict[str, object]]) -> str:
    if not rows:
        raise ValueError("Berkas tidak memiliki baris data")
    headers = {normalize_header(key) for key in rows[0]}
    if "no ticket" in headers and "pic name" in headers:
        return "FNA"
    if "ticket number" in headers and "assignee name" in headers:
        return "BBM"
    first_ticket = normalize_text(_first(rows[0], "Ticket No", "Ticket Number", "No ticket")) or ""
    prefix = first_ticket[:3].upper()
    if prefix in {"PMS", "PMG"}:
        return prefix
    name = filename.casefold()
    if "genset" in name:
        return "PMG"
    if "pm site" in name or "pm_site" in name:
        return "PMS"
    raise ValueError("Jenis ticket PMS/PMG tidak dapat dikenali dari header, nomor ticket, atau nama file")


def _parse_non_inap_file(filename: str, content: bytes) -> list[ParsedRow]:
    _, source_rows = _tabular_rows(filename, content)
    ticket_type = _classify_ticket_type(filename, source_rows)
    parsed: list[ParsedRow] = []
    for index, source in enumerate(source_rows, start=2):
        if not any(normalize_text(value) for value in source.values()):
            continue
        ticket = normalize_text(_first(source, "No ticket", "Ticket Number", "Ticket No"))
        pic = normalize_text(_first(source, "Pic name", "Assignee Name", "PIC"))
        date_value = _first(source, "Created at", "Date")
        if ticket_type == "PMG":
            date_value = _first(source, "Schedule Date", "Created Date", "Created at", "Date")
        elif ticket_type == "PMS":
            date_value = _first(source, "Created Date", "Schedule Date", "Created at", "Date")
        errors: list[str] = []
        if not ticket:
            errors.append("Nomor ticket wajib diisi")
        if any(str(value).startswith("=") for value in (ticket, pic, date_value) if value is not None):
            errors.append("Formula tidak diizinkan pada kolom kunci")
        try:
            ticket_date = _parse_date(date_value)
        except ValueError as exc:
            errors.append(str(exc))
            ticket_date = None
        payload = {
            "ticket_number": ticket,
            "ticket_type": ticket_type,
            "ticket_date": ticket_date.isoformat() if ticket_date else None,
            "status": normalize_text(_first(source, "Status", "Ticket status", "Ticket Status")),
            "site_id": normalize_text(_first(source, "Site Id", "Site ID", "Siteid")),
            "site_name": normalize_text(_first(source, "Site Name", "Sitename")),
            "nop": normalize_text(_first(source, "NOP", "Witel", "Territory")),
            "regional": normalize_text(_first(source, "Regional", "Region")),
            "pic_takeover_raw": pic,
            "pic_takeover_key": normalize_pic_key(pic),
            "source_file": filename,
            "source_row": index,
            "source_payload": {key: _json_value(value) for key, value in source.items()},
        }
        payload["source_hash"] = _source_hash(
            {key: value for key, value in payload.items() if key not in {"source_file", "source_row"}}
        )
        parsed.append(ParsedRow(filename, index, ticket, payload, errors))
    return parsed


def _nullable(value: object) -> str | None:
    return normalize_text(value)


def _fault_datetime(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    normalized = _nullable(value)
    if normalized is None:
        return None
    for fmt in (
        "%d/%m/%Y %H:%M", "%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S.%f",
    ):
        try:
            return datetime.strptime(normalized, fmt).replace(microsecond=0)
        except ValueError:
            continue
    raise ValueError(f"Timestamp tidak valid: {normalized}")


def _fault_timestamp(value: object) -> str | None:
    parsed = _fault_datetime(value)
    return parsed.isoformat(sep=" ") if parsed else None


def _fault_interval(value: object) -> str | None:
    value = _nullable(value)
    if value is None:
        return None
    if value == "0":
        return "0:00:00"
    if not re.fullmatch(r"-?\d+:\d{2}:\d{2}", value):
        raise ValueError(f"Interval tidak valid: {value}")
    return value


def _fault_integer(value: object) -> int | None:
    value = _nullable(value)
    if value is None:
        return None
    return int(value)


def _fault_decimal(value: object) -> str | None:
    value = _nullable(value)
    if value is None:
        return None
    if "," in value:
        value = value.replace(".", "").replace(",", ".")
    try:
        return str(Decimal(value))
    except InvalidOperation as exc:
        raise ValueError(f"Angka tidak valid: {value}") from exc


def _fault_boolean(value: object) -> bool | None:
    value = _nullable(value)
    if value is None:
        return None
    normalized = value.casefold()
    if normalized in {"true", "yes", "auto resolved"}:
        return True
    if normalized in {"false", "no", "manual resolved"}:
        return False
    raise ValueError(f"Boolean tidak valid: {value}")


def _fault_interval_from_seconds(seconds: float | None) -> str | None:
    if seconds is None:
        return None
    total_seconds = int(seconds)
    sign = "-" if total_seconds < 0 else ""
    hours, remainder = divmod(abs(total_seconds), 3600)
    minutes, remaining_seconds = divmod(remainder, 60)
    return f"{sign}{hours}:{minutes:02d}:{remaining_seconds:02d}"


def _fault_duration_bucket(seconds: float | None) -> str | None:
    if seconds is None:
        return None
    if seconds < 59 * 60:
        return "<1 Jam"
    if seconds < 7200:
        return "1-2 Jam"
    if seconds < 14400:
        return "2-4 Jam"
    return ">4 Jam"


def _raw_fault_source(source: dict[str, object]) -> dict[str, object]:
    normalized = {header: source.get(header) for header in FAULT_CSV_TO_COLUMN}
    for legacy_header, raw_header in RAW_FAULT_RENAMED_HEADERS.items():
        normalized[legacy_header] = source.get(raw_header)

    ticket = normalize_text(source.get("Ticket Number SWFM")) or ""
    occured_at = _fault_datetime(source.get("Occured Time"))
    created_at = _fault_datetime(source.get("Created At"))
    cleared_at = _fault_datetime(source.get("Cleared Time"))
    takeover_at = _fault_datetime(source.get("Take Over Date"))
    period_at = occured_at or created_at
    mttr_seconds = (
        (cleared_at - occured_at).total_seconds()
        if cleared_at is not None and occured_at is not None
        else None
    )
    response_seconds = (
        (takeover_at - created_at).total_seconds()
        if takeover_at is not None and created_at is not None
        else 0
    )
    check_in_at = _fault_datetime(source.get("Check In At"))
    rh_start_at = _fault_datetime(source.get("RH Start Time"))

    normalized.update({
        "Kategori TT": ticket[:3] or None,
        "Kabupaten Kota": None,
        "Tahun": period_at.year if period_at else None,
        "Periode/Bulan": month_name[period_at.month] if period_at else None,
        "Tanggal": period_at.day if period_at else None,
        "MTTR": _fault_interval_from_seconds(mttr_seconds),
        "Respon time": _fault_interval_from_seconds(response_seconds),
        "TakeOver": "TAKE OVER" if takeover_at else "NOT TAKEN",
        "PLN downtime": (
            str((Decimal(str(mttr_seconds)) / Decimal(60)).quantize(Decimal("0.01")))
            if mttr_seconds is not None else None
        ),
        "Durasi": _fault_duration_bucket(mttr_seconds),
        "Visitation": "Visit site" if check_in_at else "Not Visit",
        "Backup sukses": "BU Genset" if rh_start_at else "Not BU Genset",
    })
    return normalized


def _parse_fault_file(filename: str, content: bytes) -> ParsedImport:
    headers, source_rows = _tabular_rows(filename, content)
    legacy_missing = set(FAULT_CSV_TO_COLUMN).difference(headers)
    raw_missing = RAW_FAULT_REQUIRED_HEADERS.difference(headers)
    source_format = "legacy"
    if legacy_missing:
        if raw_missing:
            raise ValueError(
                f"Kolom wajib tidak ditemukan: {', '.join(sorted(legacy_missing)[:8])}"
            )
        source_format = "raw_swfm"
    parsed: list[ParsedRow] = []
    for index, source in enumerate(source_rows, start=2):
        if not any(normalize_text(value) for value in source.values()):
            continue
        payload: dict[str, object] = {}
        errors: list[str] = []
        if source_format == "raw_swfm":
            try:
                source = _raw_fault_source(source)
            except (ValueError, TypeError) as exc:
                errors.append(f"Normalisasi SWFM: {exc}")
        for source_column, db_column in FAULT_CSV_TO_COLUMN.items():
            value = source.get(source_column)
            try:
                if db_column in FAULT_TIMESTAMPS:
                    normalized = _fault_timestamp(value)
                elif db_column in FAULT_INTERVALS:
                    normalized = _fault_interval(value)
                elif db_column in FAULT_INTEGERS:
                    normalized = _fault_integer(value)
                elif db_column in FAULT_DECIMALS:
                    normalized = _fault_decimal(value)
                elif db_column in FAULT_BOOLEANS:
                    normalized = _fault_boolean(value)
                else:
                    normalized = _nullable(value)
                payload[db_column] = normalized
            except (ValueError, TypeError) as exc:
                errors.append(f"{source_column}: {exc}")
                payload[db_column] = None
        ticket = normalize_text(payload.get("ticket_number_swfm"))
        if not ticket:
            errors.append("Ticket Number SWFM wajib diisi")
        parsed.append(ParsedRow(filename, index, ticket, payload, errors))

    years = {row.payload.get("tahun") for row in parsed if not row.errors}
    periods = {row.payload.get("periode_bulan") for row in parsed if not row.errors}
    errors = []
    if len(years) != 1 or None in years or len(periods) != 1 or None in periods:
        errors.append("File harus berisi tepat satu tahun dan satu periode bulan")
    metadata = {
        "year": next(iter(years)) if len(years) == 1 else None,
        "period": next(iter(periods)) if len(periods) == 1 else None,
    }
    if source_format == "raw_swfm":
        metadata["source_format"] = source_format
    if errors:
        for row in parsed:
            row.errors.extend(errors)
    warnings = []
    if source_format == "raw_swfm":
        missing_mttr = sum(1 for row in parsed if row.payload.get("mttr") is None)
        if missing_mttr:
            warnings.append(
                f"{missing_mttr} baris tidak memiliki Occured Time/Cleared Time; MTTR dikosongkan"
            )
    return ParsedImport(parsed, warnings, metadata)


def _mark_duplicates(rows: list[ParsedRow]) -> None:
    counts = Counter(row.row_key for row in rows if row.row_key)
    duplicates = {key for key, count in counts.items() if count > 1}
    for row in rows:
        if row.row_key in duplicates:
            row.errors.append("Nomor ticket duplikat di dalam berkas upload")


def _fault_site_lookup_keys(row: ParsedRow) -> list[str]:
    keys: list[str] = []
    site_id = normalize_text(row.payload.get("site_id"))
    if site_id:
        keys.append(site_id.upper())
    site_name = normalize_text(row.payload.get("site_name")) or ""
    site_code = re.match(r"([A-Za-z]{3}\d{3})", site_name)
    if site_code and site_code.group(1).upper() not in keys:
        keys.append(site_code.group(1).upper())
    return keys


def _fault_location(value: object) -> str | None:
    normalized = normalize_text(value)
    if normalized is None:
        return None
    if normalized.casefold() in FAULT_MISSING_LOCATION_VALUES:
        return None
    if normalized.upper().startswith(("#N/A", "#REF!")):
        return None
    return normalized


async def _enrich_raw_fault_locations(session: object, rows: list[ParsedRow]) -> None:
    lookup_keys = sorted({key for row in rows for key in _fault_site_lookup_keys(row)})
    if not lookup_keys:
        for row in rows:
            row.errors.append("Kabupaten Kota tidak dapat dicari karena Site Id/Site Name kosong")
        return
    result = await session.execute(
        text(
            """
            SELECT UPPER(TRIM("Siteid")) AS site_id,
                   MAX(NULLIF(TRIM("Kabupaten/KOTA"), '')) AS kabupaten
            FROM data_site_master
            WHERE UPPER(TRIM("Siteid")) = ANY(CAST(:site_ids AS text[]))
            GROUP BY UPPER(TRIM("Siteid"))
            """
        ),
        {"site_ids": lookup_keys},
    )
    site_locations = {}
    for record in result.mappings():
        site_id = normalize_text(record["site_id"])
        location = _fault_location(record["kabupaten"])
        if site_id and location:
            site_locations[site_id.upper()] = location
    for row in rows:
        location = next(
            (site_locations[key] for key in _fault_site_lookup_keys(row) if key in site_locations),
            None,
        )
        row.payload["kabupaten_kota"] = location
        if location is None:
            site_id = normalize_text(row.payload.get("site_id")) or "(kosong)"
            row.errors.append(f"Kabupaten Kota tidak ditemukan untuk Site Id {site_id}")


async def validate_import(target: str, uploads: list[UploadFile], actor: AppUser) -> dict[str, object]:
    if target not in TARGETS:
        raise HTTPException(status_code=404, detail="Target import tidak tersedia")
    if not uploads:
        raise HTTPException(status_code=422, detail="Pilih minimal satu berkas")
    if target == "ticketing_fault_center" and len(uploads) != 1:
        raise HTTPException(status_code=422, detail="Fault Center hanya menerima satu berkas per import")

    files: list[tuple[str, bytes]] = []
    try:
        for upload in uploads:
            files.append(await _read_upload(upload))
        if sum(len(content) for _, content in files) > MAX_TOTAL_BYTES:
            raise ValueError("Total ukuran upload melebihi 20 MB")
        if target == "ticketing_swfm_non_inap":
            parsed = ParsedImport(
                rows=[row for filename, content in files for row in _parse_non_inap_file(filename, content)],
                warnings=[],
                metadata={"ticket_types": []},
            )
            parsed.metadata["ticket_types"] = sorted({row.payload.get("ticket_type") for row in parsed.rows})
        else:
            parsed = _parse_fault_file(*files[0])
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    if not parsed.rows:
        raise HTTPException(status_code=422, detail="Tidak ada baris data yang dapat diproses")
    _mark_duplicates(parsed.rows)
    if target == "ticketing_swfm_non_inap":
        blank_pic = sum(1 for row in parsed.rows if not row.payload.get("pic_takeover_key"))
        if blank_pic:
            parsed.warnings.append(f"{blank_pic} baris memiliki PIC kosong dan tidak masuk ranking takeover")

    existing: dict[str, str] = {}
    async with async_session() as session:
        if target == "ticketing_fault_center" and parsed.metadata.get("source_format") == "raw_swfm":
            await _enrich_raw_fault_locations(session, parsed.rows)
        valid_rows = [row for row in parsed.rows if not row.errors]
        if valid_rows:
            keys = [row.row_key for row in valid_rows]
            if target == "ticketing_swfm_non_inap":
                result = await session.execute(
                    text(
                        "SELECT ticket_number, source_hash FROM ticketing_swfm_non_inap "
                        "WHERE ticket_number = ANY(CAST(:keys AS text[]))"
                    ),
                    {"keys": keys},
                )
                existing = {row.ticket_number: row.source_hash for row in result}
            else:
                result = await session.execute(
                    text(
                        "SELECT ticket_number_swfm, tahun, periode_bulan FROM ticketing_fault_center "
                        "WHERE ticket_number_swfm = ANY(CAST(:keys AS text[]))"
                    ),
                    {"keys": keys},
                )
                existing_rows = list(result)
                existing = {row.ticket_number_swfm: "" for row in existing_rows}
                expected_year = parsed.metadata.get("year")
                expected_periods = _period_variants(str(parsed.metadata.get("period") or ""))
                collisions = {
                    row.ticket_number_swfm
                    for row in existing_rows
                    if row.tahun != expected_year
                    or (normalize_text(row.periode_bulan) or "").casefold() not in expected_periods
                }
                for parsed_row in parsed.rows:
                    if parsed_row.row_key in collisions:
                        parsed_row.errors.append("Nomor ticket sudah ada pada periode lain")

        for row in parsed.rows:
            if row.errors:
                row.change_kind = "invalid"
            elif row.row_key not in existing:
                row.change_kind = "insert"
            elif target == "ticketing_swfm_non_inap" and existing[row.row_key] == row.payload["source_hash"]:
                row.change_kind = "unchanged"
            else:
                row.change_kind = "update"

        job_id = str(uuid4())
        counts = Counter(row.change_kind for row in parsed.rows)
        await session.execute(
            text(
                """
                INSERT INTO data_import_jobs (
                    id, target, strategy, status, actor_username, file_count, source_rows,
                    valid_rows, invalid_rows, inserted_rows, updated_rows, unchanged_rows,
                    warnings, errors, metadata
                ) VALUES (
                    CAST(:id AS uuid), :target, :strategy, 'validated', :actor, :file_count,
                    :source_rows, :valid_rows, :invalid_rows, :inserted_rows, :updated_rows,
                    :unchanged_rows, CAST(:warnings AS jsonb), '[]'::jsonb, CAST(:metadata AS jsonb)
                )
                """
            ),
            {
                "id": job_id,
                "target": target,
                "strategy": TARGETS[target]["strategy"],
                "actor": actor.username,
                "file_count": len(files),
                "source_rows": len(parsed.rows),
                "valid_rows": len(parsed.rows) - counts["invalid"],
                "invalid_rows": counts["invalid"],
                "inserted_rows": counts["insert"],
                "updated_rows": counts["update"],
                "unchanged_rows": counts["unchanged"],
                "warnings": json.dumps(parsed.warnings, ensure_ascii=False),
                "metadata": json.dumps(parsed.metadata, ensure_ascii=False),
            },
        )
        stage_params = [
            {
                "job_id": job_id,
                "source_file": row.source_file,
                "source_row": row.source_row,
                "row_key": row.row_key,
                "change_kind": row.change_kind,
                "payload": json.dumps(row.payload, ensure_ascii=False),
                "validation_errors": json.dumps(row.errors, ensure_ascii=False),
            }
            for row in parsed.rows
        ]
        await session.execute(
            text(
                """
                INSERT INTO data_import_job_rows (
                    job_id, source_file, source_row, row_key, change_kind, payload, validation_errors
                ) VALUES (
                    CAST(:job_id AS uuid), :source_file, :source_row, :row_key, :change_kind,
                    CAST(:payload AS jsonb), CAST(:validation_errors AS jsonb)
                )
                """
            ),
            stage_params,
        )
        await session.commit()

    return _job_response(job_id, target, actor.username, counts, parsed, len(files))


def _job_response(
    job_id: str,
    target: str,
    actor: str,
    counts: Counter,
    parsed: ParsedImport,
    file_count: int,
) -> dict[str, object]:
    return {
        "id": job_id,
        "target": target,
        "strategy": TARGETS[target]["strategy"],
        "status": "validated",
        "actor_username": actor,
        "file_count": file_count,
        "source_rows": len(parsed.rows),
        "valid_rows": len(parsed.rows) - counts["invalid"],
        "invalid_rows": counts["invalid"],
        "inserted_rows": counts["insert"],
        "updated_rows": counts["update"],
        "unchanged_rows": counts["unchanged"],
        "warnings": parsed.warnings,
        "errors": [],
        "metadata": parsed.metadata,
        "preview_rows": [
            {
                "source_file": row.source_file,
                "source_row": row.source_row,
                "row_key": row.row_key,
                "change_kind": row.change_kind,
                "ticket_type": row.payload.get("ticket_type"),
                "ticket_date": row.payload.get("ticket_date") or row.payload.get("created_at"),
                "pic": row.payload.get("pic_takeover_raw") or row.payload.get("pic_take_over_ticket"),
                "errors": row.errors,
            }
            for row in parsed.rows[:MAX_PREVIEW_ROWS]
        ],
    }


def _period_variants(period: str) -> list[str]:
    normalized = period.casefold()
    for variants in FAULT_MONTH_VARIANTS:
        if normalized in variants:
            return sorted(variants)
    return [normalized]


def _fault_insert_statement() -> str:
    columns = list(FAULT_CSV_TO_COLUMN.values())
    expressions = []
    for column in columns:
        if column in FAULT_TIMESTAMPS:
            expressions.append(f"CAST(:{column} AS timestamp)")
        elif column in FAULT_INTERVALS:
            expressions.append(f"CAST(:{column} AS interval)")
        elif column in FAULT_INTEGERS:
            expressions.append(f"CAST(:{column} AS integer)")
        elif column in FAULT_DECIMALS:
            expressions.append(f"CAST(:{column} AS numeric)")
        elif column in FAULT_BOOLEANS:
            expressions.append(f"CAST(:{column} AS boolean)")
        else:
            expressions.append(f":{column}")
    return (
        f"INSERT INTO ticketing_fault_center ({', '.join(columns)}) "
        f"VALUES ({', '.join(expressions)})"
    )


def _prepare_non_inap_commit_rows(
    staged_rows: Iterable[Mapping[str, object]],
    job_id: str,
) -> list[dict[str, object]]:
    changed: list[dict[str, object]] = []
    for row in staged_rows:
        if row["change_kind"] == "unchanged":
            continue
        payload = dict(row["payload"])
        ticket_date = payload.get("ticket_date")
        if isinstance(ticket_date, str):
            payload["ticket_date"] = date.fromisoformat(ticket_date)
        payload["job_id"] = job_id
        payload["source_payload"] = json.dumps(payload["source_payload"], ensure_ascii=False)
        changed.append(payload)
    return changed


async def commit_import(job_id: str, actor: AppUser) -> dict[str, object]:
    async with async_session() as session:
        result = await session.execute(
            text("SELECT * FROM data_import_jobs WHERE id = CAST(:id AS uuid) FOR UPDATE"),
            {"id": job_id},
        )
        job = result.mappings().first()
        if job is None:
            raise HTTPException(status_code=404, detail="Job import tidak ditemukan")
        if job["status"] != "validated":
            raise HTTPException(status_code=409, detail="Job import tidak lagi dapat dijalankan")
        if job["invalid_rows"]:
            raise HTTPException(status_code=409, detail="Perbaiki semua baris invalid sebelum commit")
        if job["actor_username"].casefold() != actor.username.casefold() and actor.role != "sysadmin":
            raise HTTPException(status_code=403, detail="Job import dimiliki pengguna lain")

        try:
            await session.execute(
                text("SELECT pg_advisory_xact_lock(hashtext(:target))"),
                {"target": job["target"]},
            )
            await session.execute(
                text(
                    "UPDATE data_import_jobs SET status = 'committing', updated_at = NOW() "
                    "WHERE id = CAST(:id AS uuid)"
                ),
                {"id": job_id},
            )
            staged = await session.execute(
                text(
                    """
                    SELECT payload, change_kind FROM data_import_job_rows
                    WHERE job_id = CAST(:id AS uuid) AND change_kind <> 'invalid'
                    ORDER BY id
                    """
                ),
                {"id": job_id},
            )
            staged_rows = list(staged.mappings())
            if job["target"] == "ticketing_swfm_non_inap":
                changed = _prepare_non_inap_commit_rows(staged_rows, job_id)
                if changed:
                    await session.execute(
                        text(
                            """
                            INSERT INTO ticketing_swfm_non_inap (
                                ticket_number, ticket_type, ticket_date, status, site_id, site_name,
                                nop, regional, pic_takeover_raw, pic_takeover_key, source_file,
                                source_row, source_payload, source_hash, import_job_id
                            ) VALUES (
                                :ticket_number, :ticket_type, CAST(:ticket_date AS date), :status,
                                :site_id, :site_name, :nop, :regional, :pic_takeover_raw,
                                :pic_takeover_key, :source_file, :source_row,
                                CAST(:source_payload AS jsonb), :source_hash, CAST(:job_id AS uuid)
                            )
                            ON CONFLICT (ticket_number) DO UPDATE SET
                                ticket_type = EXCLUDED.ticket_type,
                                ticket_date = EXCLUDED.ticket_date,
                                status = EXCLUDED.status,
                                site_id = EXCLUDED.site_id,
                                site_name = EXCLUDED.site_name,
                                nop = EXCLUDED.nop,
                                regional = EXCLUDED.regional,
                                pic_takeover_raw = EXCLUDED.pic_takeover_raw,
                                pic_takeover_key = EXCLUDED.pic_takeover_key,
                                source_file = EXCLUDED.source_file,
                                source_row = EXCLUDED.source_row,
                                source_payload = EXCLUDED.source_payload,
                                source_hash = EXCLUDED.source_hash,
                                import_job_id = EXCLUDED.import_job_id,
                                updated_at = NOW()
                            """
                        ),
                        changed,
                    )
            else:
                metadata = dict(job["metadata"])
                period = str(metadata["period"])
                await session.execute(
                    text(
                        """
                        DELETE FROM ticketing_fault_center
                        WHERE tahun = :year
                          AND LOWER(BTRIM(COALESCE(periode_bulan, ''))) = ANY(CAST(:periods AS text[]))
                        """
                    ),
                    {"year": metadata["year"], "periods": _period_variants(period)},
                )
                fault_rows = [dict(row["payload"]) for row in staged_rows]
                if fault_rows:
                    await session.execute(text(_fault_insert_statement()), fault_rows)

            await session.execute(
                text(
                    """
                    UPDATE data_import_jobs
                    SET status = 'completed', committed_at = NOW(), updated_at = NOW()
                    WHERE id = CAST(:id AS uuid)
                    """
                ),
                {"id": job_id},
            )
            await session.commit()
        except Exception as exc:
            await session.rollback()
            async with async_session() as failure_session:
                await failure_session.execute(
                    text(
                        "UPDATE data_import_jobs SET status = 'failed', updated_at = NOW() "
                        "WHERE id = CAST(:id AS uuid) AND status <> 'completed'"
                    ),
                    {"id": job_id},
                )
                await failure_session.commit()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=(
                    "Commit import gagal. Data target tidak berubah; silakan upload ulang. "
                    "Jika masalah berulang, periksa log aplikasi."
                ),
            ) from exc
    return await get_job(job_id)


async def get_job(job_id: str) -> dict[str, object]:
    async with async_session() as session:
        result = await session.execute(
            text("SELECT * FROM data_import_jobs WHERE id = CAST(:id AS uuid)"),
            {"id": job_id},
        )
        row = result.mappings().first()
        if row is None:
            raise HTTPException(status_code=404, detail="Job import tidak ditemukan")
        return _serialize_job(row)


async def list_jobs(limit: int = 30) -> list[dict[str, object]]:
    async with async_session() as session:
        result = await session.execute(
            text("SELECT * FROM data_import_jobs ORDER BY created_at DESC LIMIT :limit"),
            {"limit": limit},
        )
        return [_serialize_job(row) for row in result.mappings()]


def _serialize_job(row: object) -> dict[str, object]:
    fields = (
        "id", "target", "strategy", "status", "actor_username", "file_count",
        "source_rows", "valid_rows", "invalid_rows", "inserted_rows", "updated_rows",
        "unchanged_rows", "warnings", "errors", "metadata", "created_at", "committed_at",
    )
    return {
        key: _json_value(row[key]) if key not in {"id", "warnings", "errors", "metadata"} else (
            str(row[key]) if key == "id" else row[key]
        )
        for key in fields
    }
