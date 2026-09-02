"""Backend-owned XLSX workbooks for Network Reporting analysis."""

from __future__ import annotations

from collections import OrderedDict
from io import BytesIO
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
COUNT_FORMAT = "#,##0"
PERCENT_FORMAT = "0.0%"
CURRENCY_FORMAT = '"Rp" #,##0'
AVAILABILITY_FORMAT = "0.00%"

TITLE_FILL = PatternFill("solid", fgColor="111827")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
STRIPE_FILL = PatternFill("solid", fgColor="F3F4F6")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
SUBTLE_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))


AREA_HEADERS = (
    "Kabupaten",
    "Site",
    "U30",
    "U30 MoM",
    "U60",
    "U60 MoM",
    "Revenue",
    "Revenue Sebelumnya",
    "Revenue MoM",
    "Payload",
    "Payload Sebelumnya",
    "Payload MoM",
    "Availability",
    "Availability MoM",
    "Ticket BPS",
    "Ticket TS",
    "Backup Success Rate",
    "Proker Open",
    "Proker Closed",
)

SITE_HEADERS = (
    "Site ID",
    "Site Name",
    "NOP",
    "Kabupaten",
    "Site Class",
    "Transport",
    "Status",
    "Revenue",
    "Revenue Sebelumnya",
    "Revenue MoM",
    "Payload",
    "Payload Sebelumnya",
    "Payload MoM",
    "Availability",
    "Availability MoM",
)

FIELD_LABELS = {
    "period": "Periode",
    "nop": "NOP",
    "kabupaten": "Kabupaten",
    "site_id": "Site ID",
    "site_class": "Site Class",
    "transport_type": "Transport",
    "mapping_status": "Mapping Status",
    "ticket_category": "Ticket Category",
    "backup_result": "Backup Result",
    "status": "Status",
    "sites": "Site",
    "revenue": "Revenue",
    "revenue_per_site": "Revenue per Site",
    "payload": "Payload",
    "payload_per_site": "Payload per Site",
    "traffic": "Traffic",
    "availability": "Availability",
    "outage_minutes": "Outage Minutes",
    "tickets": "Ticket",
    "bps_tickets": "Ticket BPS",
    "ts_tickets": "Ticket TS",
    "backup_success": "Backup Success",
    "backup_success_rate": "Backup Success Rate",
    "activities": "Activity",
    "open_activities": "Open Activity",
    "closed_activities": "Closed Activity",
}

PERCENT_FIELDS = {"availability", "backup_success_rate"}
CURRENCY_FIELDS = {"revenue", "revenue_per_site"}


def _value(item, field, default=None):
    if isinstance(item, dict):
        return item.get(field, default)
    return getattr(item, field, default)


def _percent_fraction(value):
    if value is None:
        return None
    return round(float(value) / 100.0, 10)


def _band_label(value: str | None) -> str:
    return {
        "u30": "U30",
        "u60": "U60",
        "achieved": "Achieved",
        "unavailable": "Unavailable",
    }.get(str(value or "").lower(), "Unavailable")


def _field_label(field: str) -> str:
    return FIELD_LABELS.get(field, field.replace("_", " ").title())


def _prepare_sheet(sheet, *, title: str, period: str, scope: str, headers: tuple[str, ...]):
    last_column = get_column_letter(len(headers))
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells(f"A1:{last_column}1")
    title_cell = sheet["A1"]
    title_cell.value = title
    title_cell.fill = TITLE_FILL
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 26

    sheet["A2"] = "Periode"
    sheet["B2"] = period
    sheet.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(headers))
    sheet["A3"] = "Area"
    sheet["B3"] = scope
    sheet.merge_cells(start_row=3, start_column=2, end_row=3, end_column=len(headers))
    for row_index in (2, 3):
        sheet.cell(row_index, 1).font = Font(bold=True, color="4B5563")
        sheet.cell(row_index, 2).font = Font(color="374151")

    sheet.append([])
    sheet.append(list(headers))
    for cell in sheet[5]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[5].height = 30
    sheet.freeze_panes = "A6"


def _finish_sheet(sheet, *, row_count: int, column_count: int):
    end_row = 5 + row_count
    sheet.auto_filter.ref = f"A5:{get_column_letter(column_count)}{max(5, end_row)}"
    for row_index in range(6, end_row + 1):
        if row_index % 2 == 0:
            for cell in sheet[row_index]:
                cell.fill = STRIPE_FILL
        for cell in sheet[row_index]:
            cell.border = SUBTLE_BORDER
            cell.alignment = Alignment(vertical="top")

    for column_index in range(1, column_count + 1):
        values = [sheet.cell(row, column_index).value for row in range(5, end_row + 1)]
        width = min(34, max(10, max((len(str(value)) for value in values if value is not None), default=10) + 2))
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _workbook_bytes(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_area_workbook(
    *,
    period_start: str,
    period_end: str,
    scope_label: str,
    areas,
    sites,
) -> bytes:
    workbook = Workbook()
    kabupaten_sheet = workbook.active
    kabupaten_sheet.title = "Kabupaten"
    site_sheet = workbook.create_sheet("Site")
    period_label = period_start if period_start == period_end else f"{period_start} - {period_end}"

    _prepare_sheet(
        kabupaten_sheet,
        title="Network Reporting - Kabupaten",
        period=period_label,
        scope=scope_label,
        headers=AREA_HEADERS,
    )
    for area in areas:
        kabupaten_sheet.append(
            [
                _value(area, "kabupaten"),
                int(_value(area, "total_sites", 0) or 0),
                int(_value(area, "u30_sites", 0) or 0),
                _percent_fraction(_value(area, "u30_mom_pct")),
                int(_value(area, "u60_sites", 0) or 0),
                _percent_fraction(_value(area, "u60_mom_pct")),
                int(_value(area, "revenue", 0) or 0),
                int(_value(area, "previous_revenue", 0) or 0),
                _percent_fraction(_value(area, "revenue_delta_pct")),
                int(_value(area, "payload", 0) or 0),
                int(_value(area, "previous_payload", 0) or 0),
                _percent_fraction(_value(area, "payload_delta_pct")),
                _percent_fraction(_value(area, "avg_availability")),
                _percent_fraction(_value(area, "availability_delta_pct")),
                int(_value(area, "ticket_swfm_bps", 0) or 0),
                int(_value(area, "ticket_swfm_ts", 0) or 0),
                _percent_fraction(_value(area, "backup_sukses_rate")),
                int(_value(area, "proker_open", 0) or 0),
                int(_value(area, "proker_closed", 0) or 0),
            ]
        )
    for row in range(6, 6 + len(areas)):
        for column in (2, 3, 5, 10, 11, 15, 16, 18, 19):
            kabupaten_sheet.cell(row, column).number_format = COUNT_FORMAT
        for column in (4, 6, 9, 12, 17):
            kabupaten_sheet.cell(row, column).number_format = PERCENT_FORMAT
        for column in (7, 8):
            kabupaten_sheet.cell(row, column).number_format = CURRENCY_FORMAT
        for column in (13, 14):
            kabupaten_sheet.cell(row, column).number_format = AVAILABILITY_FORMAT
    _finish_sheet(kabupaten_sheet, row_count=len(areas), column_count=len(AREA_HEADERS))

    _prepare_sheet(
        site_sheet,
        title="Network Reporting - Site",
        period=period_label,
        scope=scope_label,
        headers=SITE_HEADERS,
    )
    for site in sites:
        site_sheet.append(
            [
                _value(site, "site_id"),
                _value(site, "site_name"),
                _value(site, "nop"),
                _value(site, "kabupaten") or "Belum Terpetakan",
                _value(site, "site_class"),
                _value(site, "transport_type"),
                _band_label(_value(site, "revenue_band")),
                int(_value(site, "revenue", 0) or 0),
                int(_value(site, "previous_revenue", 0) or 0),
                _percent_fraction(_value(site, "revenue_mom_pct")),
                int(_value(site, "payload", 0) or 0),
                int(_value(site, "previous_payload", 0) or 0),
                _percent_fraction(_value(site, "payload_mom_pct")),
                _percent_fraction(_value(site, "avg_availability")),
                _percent_fraction(_value(site, "availability_delta_pct")),
            ]
        )
    for row in range(6, 6 + len(sites)):
        for column in (8, 9):
            site_sheet.cell(row, column).number_format = CURRENCY_FORMAT
        for column in (10, 13):
            site_sheet.cell(row, column).number_format = PERCENT_FORMAT
        for column in (11, 12):
            site_sheet.cell(row, column).number_format = COUNT_FORMAT
        for column in (14, 15):
            site_sheet.cell(row, column).number_format = AVAILABILITY_FORMAT
    _finish_sheet(site_sheet, row_count=len(sites), column_count=len(SITE_HEADERS))
    return _workbook_bytes(workbook)


def _pivot_matrix(result):
    row_fields = list(_value(result, "row_dimensions", []) or [])
    column_fields = list(_value(result, "column_dimensions", []) or [])
    value_fields = list(_value(result, "value_fields", []) or [])
    raw_rows = list(_value(result, "rows", []) or [])

    column_labels = sorted(
        {
            " · ".join(str((_value(row, "dimensions", {}) or {}).get(field) or "Tidak Diketahui") for field in column_fields)
            for row in raw_rows
        }
    ) if column_fields else [""]
    value_columns = [(column_label, field) for column_label in column_labels for field in value_fields]
    headers = [_field_label(field) for field in row_fields]
    headers.extend(
        f"{column_label} · {_field_label(field)}" if column_label else _field_label(field)
        for column_label, field in value_columns
    )

    grouped = OrderedDict()
    for row in raw_rows:
        dimensions = _value(row, "dimensions", {}) or {}
        values = _value(row, "values", {}) or {}
        row_key = tuple(str(dimensions.get(field) or "Tidak Diketahui") for field in row_fields)
        column_label = " · ".join(str(dimensions.get(field) or "Tidak Diketahui") for field in column_fields) if column_fields else ""
        grouped.setdefault(row_key, {})
        for field in value_fields:
            grouped[row_key][(column_label, field)] = values.get(field)

    rows = []
    for row_key in sorted(grouped):
        cells = list(row_key)
        for column_key in value_columns:
            field = column_key[1]
            value = grouped[row_key].get(column_key)
            cells.append(_percent_fraction(value) if field in PERCENT_FIELDS else value)
        rows.append(cells)
    return tuple(headers), rows, value_columns, len(row_fields)


def build_pivot_workbook(*, request, result) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pivot"
    headers, rows, value_columns, row_field_count = _pivot_matrix(result)
    period_start = _value(request, "period_start")
    period_end = _value(request, "period_end")
    period_label = period_start if period_start == period_end else f"{period_start} - {period_end}"
    scope = _value(request, "nop") or "Regional Jatim"
    _prepare_sheet(
        sheet,
        title=f"Network Reporting - Pivot {_value(request, 'dataset', '').title()}",
        period=period_label,
        scope=scope,
        headers=headers,
    )
    for row in rows:
        sheet.append(row)
    for row_index in range(6, 6 + len(rows)):
        for offset, (_, field) in enumerate(value_columns, start=row_field_count + 1):
            if field in PERCENT_FIELDS:
                sheet.cell(row_index, offset).number_format = AVAILABILITY_FORMAT
            elif field in CURRENCY_FIELDS:
                sheet.cell(row_index, offset).number_format = CURRENCY_FORMAT
            else:
                sheet.cell(row_index, offset).number_format = COUNT_FORMAT
    _finish_sheet(sheet, row_count=len(rows), column_count=len(headers))
    return _workbook_bytes(workbook)


def build_xlsx_filename(*parts: str | None) -> str:
    slug = "-".join(
        re.sub(r"[^a-z0-9]+", "-", str(part or "regional-jatim").lower()).strip("-")
        for part in parts
    )
    return f"{slug}.xlsx"
