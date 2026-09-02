from io import BytesIO
from pathlib import Path
import sys

from openpyxl import load_workbook


sys.path.insert(0, str(Path(__file__).resolve().parents[1]))


def _area_rows():
    from models.reporting import ReportingAreaRow

    return [
        ReportingAreaRow(
            area_key="SIDOARJO",
            kabupaten="SIDOARJO",
            total_sites=2,
            u30_sites=1,
            previous_u30_sites=2,
            u30_mom_pct=-50,
            u60_sites=1,
            previous_u60_sites=0,
            revenue=94_400_000_000,
            previous_revenue=88_000_000_000,
            revenue_delta_pct=7.2727,
            payload=22_439_000,
            previous_payload=21_000_000,
            payload_delta_pct=6.8524,
            avg_availability=99.84,
            availability_delta_pct=-0.03,
            sla_status="met",
            ticket_swfm_bps=2,
            ticket_swfm_ts=1,
            backup_sukses_bps=1,
            backup_sukses_rate=50,
            proker_open=1,
            proker_closed=2,
        )
    ]


def _site_rows():
    from models.reporting import ReportingSiteRow

    return [
        ReportingSiteRow(
            site_id="AAA001",
            site_name="Alpha",
            nop="NOP SIDOARJO",
            kabupaten="SIDOARJO",
            site_class="GOLD",
            transport_type="FO",
            revenue_band="u30",
            revenue=20_000_000,
            previous_revenue=25_000_000,
            revenue_mom_pct=-20,
            payload=15_728_640,
            previous_payload=14_000_000,
            payload_mom_pct=12.3474,
            avg_availability=99.68,
            availability_delta_pct=0.02,
            sla_status="met",
        )
    ]


def test_area_workbook_contains_full_typed_kabupaten_and_site_sheets():
    from services.reporting_export import build_area_workbook

    payload = build_area_workbook(
        period_start="2026-05",
        period_end="2026-08",
        scope_label="SIDOARJO",
        areas=_area_rows(),
        sites=_site_rows(),
    )
    workbook = load_workbook(BytesIO(payload), data_only=False)

    assert workbook.sheetnames == ["Kabupaten", "Site"]
    assert workbook["Kabupaten"].freeze_panes == "A6"
    assert workbook["Kabupaten"].auto_filter.ref == "A5:S6"
    assert workbook["Kabupaten"]["A6"].value == "SIDOARJO"
    assert workbook["Kabupaten"]["C6"].value == 1
    assert workbook["Kabupaten"]["D6"].value == -0.5
    assert workbook["Kabupaten"]["D6"].number_format == "0.0%"
    assert workbook["Site"]["H6"].value == 20_000_000
    assert isinstance(workbook["Site"]["H6"].value, int)
    assert workbook["Site"]["G6"].value == "U30"


def test_pivot_workbook_matches_server_aggregate_dimensions_and_values():
    from models.reporting import ReportingPivotRequest, ReportingPivotResponse, ReportingPivotRow
    from services.reporting_export import build_pivot_workbook

    request = ReportingPivotRequest(
        dataset="performance",
        period_start="2026-07",
        period_end="2026-07",
        rows=["kabupaten"],
        columns=["period"],
        values=[
            {"field": "revenue", "aggregation": "sum"},
            {"field": "availability", "aggregation": "weighted_avg"},
        ],
    )
    result = ReportingPivotResponse(
        dataset="performance",
        row_dimensions=["kabupaten"],
        column_dimensions=["period"],
        value_fields=["revenue", "availability"],
        estimated_cells=2,
        rows=[
            ReportingPivotRow(
                dimensions={"kabupaten": "SIDOARJO", "period": "2026-07"},
                values={"revenue": 94_400_000_000, "availability": 99.84},
            )
        ],
    )

    payload = build_pivot_workbook(request=request, result=result)
    workbook = load_workbook(BytesIO(payload), data_only=False)
    sheet = workbook["Pivot"]

    assert workbook.sheetnames == ["Pivot"]
    assert sheet.freeze_panes == "A6"
    assert sheet["A5"].value == "Kabupaten"
    assert sheet["B5"].value == "2026-07 · Revenue"
    assert sheet["C5"].value == "2026-07 · Availability"
    assert sheet["B6"].value == 94_400_000_000
    assert sheet["C6"].value == 0.9984
    assert sheet["C6"].number_format == "0.00%"
